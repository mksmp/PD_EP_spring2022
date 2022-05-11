import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
import pypyodbc as pyodbc
import sys, os, datetime
import start

# функция подключения к базе данных, на вход требует путь к базе данных возвращает курсор, который указывает на БД

def connect_to_DateBase(fullname_db):
    try:
        conn_string = r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + fullname_db
        conn = pyodbc.connect(conn_string)
        cursor = conn.cursor()
        #print("Подключение к базе данных")
        return cursor, conn
    except pyodbc.Error as e:
        print("Ошибка подключения к базе данных", e)


def sort_modul(date):
    buf = "Первый семестр"
    full_data = []
    for i in range(len(date)):
        date_dist = date[i]
        if date_dist[2] != buf:
            full_data += sorted(date[len(full_data):i])
            buf = date_dist[2]
    full_data += sorted(date[len(full_data):i + 1])
    return full_data


# функция делает запрос в базу данных и выводит нужные значения для дальнейшего вывода в карту
# (мудуль, дисциплина, семестр, зеты(складывая все за одну дисц)), на выходу лист из листов в каждом из которых находятся данные
def select_to_DataBase(cur, id_op):
    set = []
    sem = ["Первый", "Второй", "Третий", "Четвертый", "Пятый", "Шестой", "Седьмой", "Восьмой", "Девятый", "Десятый", "Одиннадцатый", "Двенадцатый" ]
    data = []
    buf = ""
    zet = 0.0
    j = -1
    sum_zet = 0
    for i in range(len(sem)):
        cur.execute(
            'SELECT ID_module, Discipline, Period, ZET, Block, Record_type   FROM Load WHERE Period LIKE ? AND ID_OP = ?',
            [(sem[i] + " семестр"), id_op])
        for row in cur.fetchall():
            if buf != row[1]:
                buf = row[1]
                data.append(str(row[4])[:7] + " " + str(row[0]))
                data.append(row[1])
                data.append(row[2])
                set.append(data.copy())
                data_rev = set[j]
                if data_rev[1] == "Элективные курсы по физической культуре и спорту" or data_rev[1] == "Элективные дисциплины по физической культуре и спорту":
                    zet = 0
                if len(data_rev) == 3:
                    data_rev.append(int(round(zet)))
                    set[j] = data_rev.copy()
                else:
                    data_rev[3] = int(round(zet))
                    set[j] = data_rev.copy()
                sum_zet += zet
                zet = 0.0
                j += 1
            if row[3] != None and row[5] != "Факультативная":
                zet += float(row[3])
            data.clear()
    data_rev = set[-1]
    data_rev.append(int(zet))
    set[-1] = data_rev.copy()
    set = sort_modul(set)
    return set


def select_color(cur, modul):
    cur.execute('SELECT Color  FROM Module_reference WHERE ID_module LIKE ?', [modul])
    for row in cur.fetchall():
        return (row[0])

def  create_directory_of_modul(ws, modul, cur):
    adr_cell = "B"
    row = 50
    modul = list(modul)
    for i in range(len(modul)):
        dip = adr_cell + str(row) + ':' + adr_cell + str(row + 1)
        cur.execute(
            'SELECT Name_module  FROM Module_reference WHERE ID_module LIKE ?', [modul[i]])
        for r in cur.fetchall():
            modul_buf = (r[0])
        ws[adr_cell + str(row)] = modul_buf
        ws[adr_cell + str(row)].style = 'standart'
        cell = ws[adr_cell + str(row)]
        color = select_color(cur, modul[i])
        cell.fill = openpyxl.styles.PatternFill(start_color=str(color), end_color=str(color), fill_type='solid')
        ws.merge_cells(dip)
        row +=2


# функция создает карту и задаем все данные кроме предметов в семестрах, на вход требует имя карты
def CreateMap(filename_map):
    wk = xlsxwriter.Workbook(filename_map)
    ws = wk.add_worksheet()
    ws.set_column(1, 40, 40)
    wk.close()
    workbook = openpyxl.load_workbook(filename_map)
    worksheet = workbook.active
    ns = NamedStyle(name='standart')
    ns.font = Font(bold=False, size=12)
    border = Side(style='medium', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    ns.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    workbook.add_named_style(ns)
    worksheet.row_dimensions[1].height = 50
    worksheet.row_dimensions[2].height = 20
    worksheet["A2"] = "З.Е."
    worksheet['A2'].style = 'standart'
    for col in range(ord('B'), ord('C')):
        worksheet[chr(col) + str(2)] = str(col - 65) + " семестр"
        worksheet[chr(col) + str(2)].style = 'standart'
    return worksheet, workbook


# заполняем данные, размер и цвет  в ячейках карты,
# Так же мы красим предметы в соответствии с модулем
def filling_map(fullname_db, filename_map, name_map):
    cur, conn = connect_to_DateBase(fullname_db)
    cur.execute('SELECT ID_OP  FROM OP WHERE Name_OP LIKE ?', [name_map])
    id_op = cur.fetchall()[0][0]
    date = select_to_DataBase(cur, id_op)
    ws, wk = CreateMap(filename_map)
    adr_cell = "B"
    buf = "Первый семестр"
    row = 3
    i = -1
    modul = set()
    max_row = 0
    sum_row = 0
    while i < len(date) - 1:
        i += 1
        date_dist = date[i]
        if date_dist[2] == buf and date_dist[3] != 0:
            ws["A" + str(row)] = row - 2
            ws["A" + str(row)].style = 'standart'
            modul.add(str(date_dist[0])[8:])
            dip = adr_cell + str(row) + ':' + adr_cell + str(row + date_dist[3] - 1)
            ws[adr_cell + str(row)].style = 'standart'
            ws[adr_cell + str(row)] = date_dist[1]
            cell = ws[adr_cell + str(row)]
            color = select_color(cur, str(date_dist[0])[8:])
            cell.fill = openpyxl.styles.PatternFill(start_color=str(color), end_color=str(color), fill_type='solid')
            ws.merge_cells(dip)
            row += date_dist[3] - 1
            buf = date_dist[2]
            row += 1
            max_row = max(max_row, row)
        elif date_dist[3] != 0:
            adr_cell = chr(ord(adr_cell) + 1)
            ws[adr_cell + str(2)] = str(ord(adr_cell) - 65) + " семестр"
            ws[adr_cell + str(2)].style = 'standart'
            buf = date_dist[2]
            sum_row += row -3
            row = 3
            i -= 1
    ws.merge_cells('A1:' + adr_cell +'1')
    ws['A1'] = 'КАРТА ДИСЦИПЛИН'
    ws['A1'].style = 'standart'
    ws['A1'].font = Font(bold=True, size=12)
    for col in range(3, max_row):
        ws["A" + str(col)] = col - 2
        ws["A" + str(col)].style = 'standart'
        ws.row_dimensions[col].height = 25
    create_directory_of_modul(ws, modul, cur)
    wk.save(filename=filename_map)

    cur.close()
    del cur
    conn.close()
    #print("Отключение от базы данных")

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def get_file_list(name_dir):
    extensions = [".xlsx"]
    file_list = []
    for root, directories, filenames in os.walk(name_dir):
        for filename in filenames:
            if any(ext in filename for ext in extensions):
                file_list.append(os.path.join(root, filename))
    return file_list


# основная функция-связующая все части и вводит основные параметры всего
def main():
    try:
        name_dir = input("Введите название папки в которой находятся выгрузки из 1С: ")
        list_name_img = get_file_list(name_dir)
        fullname_db = resource_path('db.accdb')
        for file in list_name_img:
            start.start(file, fullname_db)
            day_time = datetime.datetime.now()
            day_time = " от " + str(day_time)[:16].replace("-", ".").replace(":", "-")
            filename_map = './Карты/КД ' + os.path.basename(file)[0:-5] + day_time + '.xlsx'
            filling_map(fullname_db, filename_map, os.path.basename(file)[0:-5])
            print("Создана карта для выгрузки ", os.path.basename(file))
        print('Программа успешно завершила свою работу!')
        input("Нажмите любую кнопку для закрытия окна программы!")
    except Exception as es:
        print(es)
        input("Нажмите любую кнопку для закрытия окна программы!")

if __name__ == "__main__":
    main()
